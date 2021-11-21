from django.shortcuts import render,HttpResponse
from django.core.files.storage import FileSystemStorage
import csv
import pandas as pd
import xlsxwriter
import shutil
import openpyxl
from openpyxl.styles import Border, Side
import os
import time




# Create your views here.
def index(request):
    return render(request,'index.html')
def submit(request):
   
    args = {}
    args1= {}
    fs=FileSystemStorage()
    # t_end = time.time() + 5
    # while time.time() < t_end:
    #     return render(request,'wait.html') 
 
    if request.method=="POST":
        file1=request.FILES['file1'] #response csv
        file2=request.FILES['file2'] #master roll no csv file
        df = pd.read_csv(file1) #response csv file 
        dict = {}
        for i in range(len(df)):            
            d = {}
            d['Timestamp'] = df.iloc[i,0]
            d['Email address'] = df.iloc[i,1]
            d['Score'] = df.iloc[i,2]
            d['Name'] = df.iloc[i,3]
            d['IITP webmail'] = df.iloc[i,4]
            d['Phone (10 digit only)'] = df.iloc[i,5]
            d['Roll Number'] = df.iloc[i,6]
            d['Answer'] = list(df.iloc[i,7:])
        
            dict[df.iloc[i,6]] = d

        if(dict.get('ANSWER') == None):
            print("No roll number with ANSWER is present, Cannot Process!")
            return render(request,'no_answers_sheet.html')
            exit()

        ans_key = dict['ANSWER']['Answer'] 
        # print(ans_key)
       # print(dict)

        df = pd.read_csv(file2)   #master Roll no csv
        for i in range(len(df)):
            if df.iloc[i,0] not in dict:
                d = {}
                d['Name'] = df.iloc[i,1]
                d['Roll Number'] = df.iloc[i,0]
                d['Answer'] = [float("nan")]*len(ans_key)
                
                dict[df.iloc[i,0]] = d

        statusAns_dict = {}
        score_after_negative_dict = {}
        for x,y in dict.items():
            correct = 0
            incorrect = 0
            na = 0
            for i in range (len(ans_key)):
                if(str(y['Answer'][i])=='nan'):
                    na+=1
                elif(y['Answer'][i]== ans_key[i]):
                    correct +=1
                else:
                    incorrect+=1

            dict[x]['correct'] = correct
            dict[x]['incorrect'] = incorrect
            dict[x]['na'] = na

            ###########################################################
            # Making 2 dictionaries to store score_after_negative and statusAns
            # Check last few lines.
            ###########################################################
            statusAns_dict[x] = '['+str(correct)+','+str(incorrect)+','+str(na)+']'
            score_after_negative_dict[x] = str(int(correct*5) + int(incorrect*-1)) + '/140'

        # directory = 'sample_output_1'
        # out_path = os.path.join('./', directory)
        # try:
        #     shutil.rmtree(out_path)
        # except OSError as e:
        #     pass
        # os.mkdir(out_path)
        # # os.remove("test.xlsx") 

        wb = openpyxl.Workbook()
        wb.create_sheet(index = 0, title = 'quiz')
        sheet = wb['quiz']
        img = openpyxl.drawing.image.Image(r'C:\Users\Asus\Desktop\project1\proj1\proj1\logo.jpeg')
        img.anchor = 'A1'
        sheet.add_image(img)

        border = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))

        sheet.column_dimensions['A'].width = 16.89
        sheet.column_dimensions['B'].width = 16.89
        sheet.column_dimensions['C'].width = 16.89
        sheet.column_dimensions['D'].width = 16.89
        sheet.column_dimensions['E'].width = 16.89

        sheet.row_dimensions[5].height = 22.8
        sheet.merge_cells('A5:E5')
        sheet['A5'].font = openpyxl.styles.Font(name = 'Century',size=18, bold=True,underline='single')
        sheet['A5'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')
        sheet['A5'] = 'Mark Sheet'

        for key in dict:
            sheet['A6'] = 'Name:'
            sheet['A6'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none')
            sheet['A6'].alignment = openpyxl.styles.Alignment(horizontal='right',vertical='bottom')

            sheet['B6'] = dict[key]['Name']
            sheet['B6'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=True,underline='none')
            sheet['B6'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='bottom')

            sheet['D6'] = 'Exam:'
            sheet['D6'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none')
            sheet['D6'].alignment = openpyxl.styles.Alignment(horizontal='right',vertical='bottom')

            sheet['E6'] = 'quiz'
            sheet['E6'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=True,underline='none')
            sheet['E6'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='bottom')

            sheet['A7'] = 'Roll Number:'
            sheet['A7'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none')
            sheet['A7'].alignment = openpyxl.styles.Alignment(horizontal='right',vertical='bottom')

            sheet['B7'] = str(key)
            sheet['B7'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=True,underline='none')
            sheet['B7'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='bottom')

            sheet['B9'] = 'Right'
            sheet['B9'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=True,underline='none')
            sheet['B9'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

            sheet['C9'] = 'Wrong'
            sheet['C9'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=True,underline='none')
            sheet['C9'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

            sheet['D9'] = 'Not Attempt'
            sheet['D9'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=True,underline='none')
            sheet['D9'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

            sheet['E9'] = 'Max'
            sheet['E9'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=True,underline='none')
            sheet['E9'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

            sheet['A10'] = 'No.'
            sheet['A10'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=True,underline='none')
            sheet['A10'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

            sheet['A11'] = 'Marking'
            sheet['A11'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=True,underline='none')
            sheet['A11'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

            sheet['A12'] = 'Total'
            sheet['A12'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=True,underline='none')
            sheet['A12'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

            sheet['B10'] = dict[key]['correct']
            sheet['B10'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none',color='008000')
            sheet['B10'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

            sheet['B11'] = '5'
            sheet['B11'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none',color='008000')
            sheet['B11'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

            sheet['B12'] = str(int(dict[key]['correct'])*5)
            sheet['B12'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none',color='008000')
            sheet['B12'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

            sheet['C10'] = dict[key]['incorrect']
            sheet['C10'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none',color='FF0000')
            sheet['C10'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

            sheet['C11'] = '-1'
            sheet['C11'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none',color='FF0000')
            sheet['C11'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

            sheet['C12'] = str(int(dict[key]['incorrect'])*-1)
            sheet['C12'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none',color='FF0000')
            sheet['C12'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

            sheet['D10'] = dict[key]['na']
            sheet['D10'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none')
            sheet['D10'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

            sheet['D11'] = '0'
            sheet['D11'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none')
            sheet['D11'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

            sheet['E10'] = '28'  #number of questions or max attempted in class??
            sheet['E10'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none')
            sheet['E10'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

            sheet['E12'] = str(int(dict[key]['correct'])*5 - int(dict[key]['incorrect'])*-1) + '/140'
            sheet['E12'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none',color='0000FF')
            sheet['E12'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

            sheet['A15'] = 'Student Ans'
            sheet['A15'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=True,underline='none')
            sheet['A15'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')
            sheet['A15'].border = border

            sheet['D15'] = 'Student Ans'
            sheet['D15'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=True,underline='none')
            sheet['D15'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')
            sheet['D15'].border = border

            sheet['B15'] = 'Correct Ans'
            sheet['B15'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=True,underline='none')
            sheet['B15'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')
            sheet['B15'].border = border

            sheet['E15'] = 'Correct Ans'
            sheet['E15'].font = openpyxl.styles.Font(name = 'Century',size=12, bold=True,underline='none')
            sheet['E15'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')
            sheet['E15'].border = border

            for r in range(9,13):
                for c in range(1,6):
                    sheet.cell(row = r , column = c).border = border

            for i in range(len(ans_key)):
                if(i+16>40):
                    sheet['E'+str(i-9)].border = border
                    
                    sheet['E'+str(i-9)] = ans_key[i]
                    sheet['E'+str(i-9)].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none',color='0000FF')
                    sheet['E'+str(i-9)].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')
                else:   
                    sheet['B'+str(i+16)].border = border
                    
                    sheet['B'+str(i+16)] = ans_key[i]
                    sheet['B'+str(i+16)].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none',color='0000FF')
                    sheet['B'+str(i+16)].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')
            
            for i in range(len(ans_key)):
                if(i+16>40):
                    sheet['D'+str(i-9)].border = border
                    
                    if(str(dict[key]['Answer'][i]) == 'nan'):
                        pass
                    elif(dict[key]['Answer'][i] == ans_key[i]):
                        sheet['D'+str(i-9)] = dict[key]['Answer'][i]
                        sheet['D'+str(i-9)].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none',color='008000')
                        sheet['D'+str(i-9)].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')
                    else:
                        sheet['D'+str(i-9)] = dict[key]['Answer'][i]
                        sheet['D'+str(i-9)].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none',color='FF0000')
                        sheet['D'+str(i-9)].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')            
                        
                else:
                    sheet['A'+str(i+16)].border = border
                        
                    if(str(dict[key]['Answer'][i]) == 'nan'):
                        pass
                    elif(dict[key]['Answer'][i] == ans_key[i]):
                        sheet['A'+str(i+16)] = dict[key]['Answer'][i]
                        sheet['A'+str(i+16)].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none',color='008000')
                        sheet['A'+str(i+16)].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')
                    else:
                        sheet['A'+str(i+16)] = dict[key]['Answer'][i]
                        sheet['A'+str(i+16)].font = openpyxl.styles.Font(name = 'Century',size=12, bold=False, underline='none',color='FF0000')
                        sheet['A'+str(i+16)].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='bottom')

            # sheetDelete = wb["Sheet"]
            # wb.remove(sheetDelete)
            print(str(key))
            fn=wb.save(r'./MyFiles'+'/' +str(key)+'.xlsx')    
            args1[str(key)] =fs.url(fn)
            args1[str(key)]=args1[str(key)]+'/{}.xlsx'.format(str(key))
           #txt3 = "My name is {}, I'm {}".format("John",36)
           # fs.save(str(key)+'.xlsx',wb.save(str(key)+'.xlsx'))
            #    C:\Users\Asus\Desktop\project1\proj1\app1\views.py
            #    C:\Users\Asus\Desktop\project1\proj1\MyFiles
            #C:\Users\Asus\Desktop\project1\proj1\templates

        ##################################################################
        #Creating concise marksheet
        ##################################################################

        # concise_df = pd.read_csv(file1)
        # concise_df.insert(6,'Score_After_Negative','bleh')
        # concise_df.insert(36, 'statusAns','bleh')

        # concise_df.set_index('Roll Number')
        # concise_df["Score_After_Negative"] = concise_df["Roll Number"].map(score_after_negative_dict)
        # concise_df["statusAns"] = concise_df["Roll Number"].map(statusAns_dict)

        # concise_df.rename(columns={'Score': 'Google_Score'}, inplace=True)
        # concise_df.to_csv('concise_marksheet.csv', index = False)
       # concise_df.to_csv(out_path+'/concise_marksheet.csv', index = False)
        
    
    fname1=fs.save(file1.name,file1)
    fname2=fs.save(file2.name,file2)   
    args1['master'] =fs.url(fname1)
    args1['response']=fs.url(fname2)   
    final_args={}
    final_args["roll_no"]=args1
    return render(request,'submit.html',final_args)    
